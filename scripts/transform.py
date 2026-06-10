# -*- coding: utf-8 -*-
"""
플레이영 고객정보 엑셀 -> Supabase 임포트용 CSV 변환 스크립트

규칙(최종 기획안 기준):
  1) 같은 전화번호의 여러 행 스탬프 값을 모두 합산한다.
  2) 합산 스탬프를 10으로 나눠, 몫(//10)을 쿠폰으로 전환, 나머지(%10)를 스탬프 잔여값으로 둔다.
  3) 엑셀 '쿠폰명' 컬럼이 한 행이라도 채워져 있으면 보유 쿠폰 +1 (종류 구분 없이 하나로 통합).
  4) 사용한 쿠폰개수는 0으로 시작.
  5) 전화번호는 숫자만 남겨 문자열로 정규화(앞자리 0 보존).

출력: data/customers.csv  (phone, stamps, coupons, coupons_used, name, memo)
"""
import csv
import os
import re
import sys
from collections import defaultdict

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else r"F:\플레이영_2023\20260430_플레이영 용인구갈점_고객정보 (1) (1).xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "customers.csv")

COL_NAME = 1      # 이름
COL_PHONE = 2     # 전화번호
COL_STAMP = 5     # 스탬프 잔여값
COL_COUPON = 7    # 쿠폰명


def norm_phone(v):
    if v is None:
        return ""
    digits = re.sub(r"\D", "", str(v))
    return digits


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    agg = defaultdict(lambda: {"stamp_sum": 0, "has_coupon": False, "name": ""})

    skipped = 0
    for r in range(2, ws.max_row + 1):
        phone = norm_phone(ws.cell(row=r, column=COL_PHONE).value)
        if not phone:
            skipped += 1
            continue
        stamp = ws.cell(row=r, column=COL_STAMP).value
        coupon = ws.cell(row=r, column=COL_COUPON).value
        name = ws.cell(row=r, column=COL_NAME).value

        rec = agg[phone]
        if isinstance(stamp, (int, float)):
            rec["stamp_sum"] += int(stamp)
        if coupon not in (None, ""):
            rec["has_coupon"] = True
        if name not in (None, "") and not rec["name"]:
            rec["name"] = str(name).strip()

    rows = []
    total_coupons = 0
    total_stamps = 0
    for phone, rec in agg.items():
        s = rec["stamp_sum"]
        coupons = s // 10 + (1 if rec["has_coupon"] else 0)
        stamps = s % 10
        rows.append({
            "phone": phone,
            "stamps": stamps,
            "coupons": coupons,
            "coupons_used": 0,
            "name": rec["name"],
            "memo": "",
        })
        total_coupons += coupons
        total_stamps += stamps

    rows.sort(key=lambda x: x["phone"])

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["phone", "stamps", "coupons", "coupons_used", "name", "memo"])
        w.writeheader()
        w.writerows(rows)

    print("unique customers:", len(rows))
    print("rows skipped (no phone):", skipped)
    print("total coupons after migration:", total_coupons)
    print("customers with >=1 coupon:", sum(1 for r in rows if r["coupons"] > 0))
    print("total remaining stamps:", total_stamps)
    print("output:", os.path.abspath(OUT))


if __name__ == "__main__":
    main()
